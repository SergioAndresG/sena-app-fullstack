import uuid
import shutil
from pathlib import Path
from typing import Optional
import hashlib
from datetime import datetime
from MODELS.archivo_excel import ArchivoExcel
from connection import SessionLocal
from sqlalchemy.orm import Session
from MODELS.ficha import Ficha
import tempfile
import base64
import asyncio
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from copy import deepcopy
from io import BytesIO
from MODELS import ArchivoExcel, Ficha
from fastapi import Depends

def capitalizar(texto: str) -> str:
    if not texto:
        return ""
    return texto.strip().lower().title()

class FormatoService:
    def __init__(self,base_path = "archivos_exportados"):
        self.base_path = Path(base_path)
        self.base_path.mkdir(parents=True, exist_ok=True)
        try:
            # Cargas tu plantilla GRUPAL
            self.plantilla_grupal_wb = Path("GRUPAL-F165.xlsx")
            # Cargas tu plantilla INDIVIDUAL
            self.plantilla_individual_wb = Path("INDIVIDUAL-F165.xlsx")
        except FileNotFoundError as e:
            # Si el archivo no existe, la aplicación no puede funcionar.
            # Es mejor lanzar un error claro aquí.
            raise RuntimeError(f"Error crítico: No se pudo encontrar el archivo de plantilla: {e.filename}")

    def calcular_hash(self, ruta_archivo: Path) -> str:
        """Calcula el hash SHA256 de un archivo."""
        sha256 = hashlib.sha256() # Crea un objeto hash SHA256
        # Abre el archivo en modo binario y lee en bloques para evitar problemas de memoria
        with ruta_archivo.open("rb") as f:
            # Lee el archivo en bloques de 8192 bytes
            # Esto es eficiente para archivos grandes
            while chunk := f.read(8192):
                # Actualiza el hash con el bloque leído
                sha256.update(chunk)
        # Devuelve el hash en formato hexadecimal
        return sha256.hexdigest()
    
    
    def generar_nombre_interno(self, extension:str="xlsx") -> str:
        """Genera un nombre interno único para el archivo."""
        timpestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        uuid_corto = str(uuid.uuid4())[:8]  # Genera un UUID y toma los primeros 8 caracteres
        return f"{uuid_corto}_{timpestamp}.{extension}"
    


    def obtener_ruta_organizada(self, nombre_interno:str) -> Path:
        """Oraganiza los archivos por año/mes para mejor gestion"""
        ahora = datetime.now()
        ruta = self.base_path / str(ahora.year) / str(ahora.month) / "exportados"
        ruta.parent.mkdir(parents=True, exist_ok=True)
        try:
            ruta.mkdir(parents=True, exist_ok=True)
            print(f"Carpetas verificadas/creadas: {ruta}")
        except Exception as e:
            print(f"Error creando carpetas en obtener_ruta_organizada: {e}")
            raise
        return ruta / f"{nombre_interno}"
    


    def guardar_archivo_seguro(self, contenido: bytes, nombre_original:str, 
                            ficha: str, modalidad: str, cantidad_aprendices:int,
                            usuario_id: Optional[int] = None, aprendiz_documento=None) -> ArchivoExcel:
        """ Guarda un archivo de Excel de manera segura con validaciones """

        try:
            #Paso 1: generar nombres y rutas
            nombre_interno = self.generar_nombre_interno()
            ruta_completa = self.obtener_ruta_organizada(nombre_interno)
            ruta_relativa = ruta_completa.relative_to(self.base_path)

            #Paso 2: escribir el archivo
            with open(ruta_completa, "wb") as f:
                f.write(contenido)

            #Paso 3: calcular el hash y tamaño
            hash_archivo = self.calcular_hash(ruta_completa)  # ← SIN str()
            tamaño_bytes = ruta_completa.stat().st_size

            #Paso 4: Crear registro en la base de datos
            archivo_db = ArchivoExcel(
                nombre_original=nombre_original,
                nombre_interno=nombre_interno,
                ruta_archivo=str(ruta_relativa),
                ficha=ficha,
                modalidad=modalidad,
                cantidad_aprendices=cantidad_aprendices,
                hash_archivo=hash_archivo,
                tamaño_bytes=tamaño_bytes,
                usuario_id=usuario_id if usuario_id else 0,
                aprendiz_documento=aprendiz_documento,
            )
            return archivo_db
        except Exception as e:
            if ruta_completa.exists():
                ruta_completa.unlink()
            raise Exception(f"Error al guardar el archivo: {str(e)}") from e
        

    def verificar_integridad_archivo(self, archivo_db: ArchivoExcel) -> bool:
        """"Verifica que el archivo no este corrupto"""
        try:
            ruta_completa = self.base_path / archivo_db.ruta_archivo
            if not ruta_completa.exists():
                return False
            hash_calculado = self.calcular_hash(str(ruta_completa))
            return hash_calculado == archivo_db.hash_archivo
        except Exception:
            return False
        
    def obtene_archivo_para_descarga(self, archivo_db: ArchivoExcel) -> Path:
        """Obtiene la ruta completa del archivo para descarga"""
        try:
            ruta_completa = self.base_path / archivo_db.ruta_archivo

            if not self.verificar_integridad_archivo(archivo_db):
                raise FileNotFoundError(f"El archivo esta corrupto o no existe: {archivo_db.nombre_interno}")
            
            with open(ruta_completa, "rb") as f:
                contenido = f.read()
        
        except Exception as e:
            raise Exception(f"Error al leer el archivo: {str(e)}") from e
        

    def eliminar_archivo_seguro(self,archivo_db: ArchivoExcel) -> bool:
        """Eliminacion segura (soft delete)"""
        try:
            # 1. Soft delete en DB
            archivo_db.activo = False
            archivo_db.fecha_modificacion = datetime.now()
            SessionLocal.commit()  # Asumiendo que tienes una sesión de DB activa

            return True
        except Exception as e:
            SessionLocal.rollback()  # Revertir cambios en caso de error
            raise Exception(f"Error al eliminar el archivo: {str(e)}") from e
        

    # Funcion que valida que encontro la ficha
    def _validar_y_obtener_ficha(self, numero_ficha: str, db:Session) -> Ficha: # <- Retornamos un objeto ficha
        try:
            ficha = db.query(Ficha).filter(Ficha.numero_ficha == numero_ficha).first()
            if ficha:
                return ficha
            else:
                raise ("Ficha no encontrada")

        except Exception as e:
            raise Exception(f"Error al obtener ficha {e}")
    
    @staticmethod
    def _procesar_imagen_individual(firma_data:str)-> str:
        try:
            #Eliminar encabezado si existe
            if "," in firma_data: 
                firma_data = firma_data.split(",")[1]
            # Decadificar base64
            firma_bytes = base64.b64decode(firma_data)

            #Crear un archivo temporal
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_file:
                tmp_file.write(firma_bytes)
                return tmp_file.name
        except Exception as e:
            print(f"Errrp procesando imagen: {e}")
            return None
        
    async def procesar_firmas_en_paralelo(self,aprendices:list) ->list:
        # Procesar las firmas en paralelo
        executor = ThreadPoolExecutor(max_workers=4)
        loop = asyncio.get_event_loop()
        
        tareas = []

        for ap in aprendices:
            if ap.firma:
                tarea = loop.run_in_executor(
                    executor,
                    FormatoService._procesar_imagen_individual,
                    ap.firma
                )
                tareas.append(tarea)
            else:
                future = loop.create_future()
                future.set_result(None)
                tareas.append(future)

        ruta_imagenes = await asyncio.gather(*tareas)
        executor.shutdown(wait=False)

        return ruta_imagenes
    

    def _llenar_F165_grupal(self,wb,ficha,aprendices,imagenes_procesadas,request,usuario_gene,informacion_adicional):

        fecha_inicio = ficha.fecha_inicio.strftime("%d-%m-%Y") if ficha.fecha_inicio else "N/A"
        fecha_fin = ficha.fecha_fin.strftime("%d-%m-%Y") if ficha.fecha_fin else "N/A"
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        
        fecha_productiva = informacion_adicional.fecha_inicio_etapa_productiva
        if isinstance(fecha_productiva, str):
            try:
                fecha_productiva = datetime.strptime(fecha_productiva, "%Y-%m-%d").date()
            except ValueError:
                fecha_productiva = None

        hoja = wb["Selección formato - Grupal"]
        hoja["E8"] = "x"
        
        hoja["E12"] = "25 / Cundinamarca"
        hoja["H12"] = "9512 / Centro De Biotecnología Agropecuaria"
        
        datos_fechas = {
            "E13": fecha_inicio,
            "H13": fecha_fin,
            "H14": fecha_fin, 
            "E11": fecha_actual,
            "E14": fecha_productiva
        }

        # Aplicar todo de una vez
        font_style = Font(size=12, bold=True)
        for celda, valor in datos_fechas.items():
            if valor:  # Si hay fecha válida
                hoja[celda].value = valor
                hoja[celda].number_format = "DD-MM-YYYY"  # <-- Forzar formato de fecha
            else:
                hoja[celda].value = "N/A"
            hoja[celda].font = font_style

        hoja["T11"] = request.ficha  # Número de ficha

        nombre_completo_instructor = capitalizar(f"{usuario_gene.nombre} {usuario_gene.apellidos}")

        correo_instructor = usuario_gene.correo
        
        hoja["T13"] = capitalizar(nombre_completo_instructor)  # Instructor
        hoja["U14"] = correo_instructor  # Correo

        hoja["J13"] = capitalizar(informacion_adicional.trimestre)
        hoja["J14"] = capitalizar(informacion_adicional.jornada)
        hoja["U12"] = capitalizar(informacion_adicional.modalidad_formacion)

        hoja["H11"] = capitalizar(f"{informacion_adicional.nivel_formacion} {ficha.programa}")
        
        fila_inicial = 18 # Los datos empiezan en la fila 18
        espacios_disponibles = 20 # La plantilla tiene 20 espacios

        aprendices_extra = len(aprendices) - espacios_disponibles if len(aprendices) > espacios_disponibles else 0

        # Inserta filas si hay más aprendices que espacios
        if aprendices_extra > 0:
            print(f"{aprendices_extra} aprendices extra")
            try:
                hoja.insert_rows(idx=fila_inicial + espacios_disponibles, amount=aprendices_extra)
                print("Filas insertadas")
            except Exception as e:
                print(f"Error al insertar filas: {e}")

        # Llenar datos de forma optimizada
        for i, (ap, imagen_path) in enumerate(zip(aprendices, imagenes_procesadas)):
            fila = fila_inicial + i
            
            # Escribir todos los datos de texto de una vez
            datos_fila = [
                (f"C{fila}", ap.tipo_documento),
                (f"D{fila}", ap.documento),
                (f"E{fila}", ap.nombre),
                (f"F{fila}", ap.apellido),
                (f"G{fila}", ap.direccion),
                (f"H{fila}", ap.correo),
                (f"I{fila}", ap.celular),
                (f"L{fila}", ap.tipo_discapacidad),
                (f"M{fila}", "x")
            ]
            
            for celda, valor in datos_fila:
                hoja[celda] = valor
            
            # Manejar discapacidad
            if ap.discapacidad == 'NO':
                hoja[f"K{fila}"] = "x"
            else:
                hoja[f"J{fila}"] = "x"

            hoja.row_dimensions[fila].height = 50
            

            if imagen_path:
                try:
                    img = OpenpyxlImage(imagen_path)
                    img.width = 120
                    img.height = 50
                    celda = f"AG{fila}"
                    hoja.add_image(img, celda)
                except Exception as e:
                    print(f"Error insertando imagen en fila {fila}: {e}")
                    print(f"Tipo de error {type(e)}")


    def _llenar_F165_individual(self,wb,ficha,aprendices,imagenes_procesadas,request,usuario_gene,informacion_adicional):
        if not aprendices:
            raise ValueError("La lista de aprendices no puede estar vacía para el formato individual.")
        fecha_inicio = ficha.fecha_inicio.strftime("%d-%m-%Y") if ficha.fecha_inicio else "N/A"
        fecha_fin = ficha.fecha_fin.strftime("%d-%m-%Y") if ficha.fecha_fin else "N/A"
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        
        fecha_productiva = informacion_adicional.fecha_inicio_etapa_productiva
        if isinstance(fecha_productiva, str):
            try:
                fecha_productiva = datetime.strptime(fecha_productiva, "%Y-%m-%d").date()
            except ValueError:
                fecha_productiva = None

        hoja = wb["Selección Modificación Indiv"]
        hoja["C8"] = "x"

        # Tomar el primer (y único) aprendiz
        ap = aprendices[0]
        imagen_path = imagenes_procesadas[0] if imagenes_procesadas else None

        # Datos del aprendiz
        hoja["C12"] = ap.tipo_documento
        hoja["D12"] = ap.documento
        hoja["E12"] = capitalizar(f"{ap.nombre} {ap.apellido}")
        hoja["F12"] = ap.celular
        hoja["G12"] = ap.correo
        hoja["B14"] = capitalizar(ap.direccion)
        hoja["C14"] = capitalizar(ap.departamento)
        hoja["D14"] = capitalizar(ap.municipio)

        if ap.discapacidad == 'NO':
            hoja["E14"] = "Si   (  )   No   ( X )"
        else:
            hoja["E14"] = "Si   ( X )   No   (  )"  

        hoja["G14"] = capitalizar(ap.tipo_discapacidad)
        hoja["B17"] = "25 / Cundinamarca"
        hoja["C17"] = "9512 / Centro De Biotecnología Agropecuaria"
        hoja["E17"] = request.ficha
        hoja["F17"] = capitalizar(informacion_adicional.nivel_formacion)
        hoja["G17"] = capitalizar(ficha.programa)
        hoja["E19"] = "Selección de alternativa: ( X )"
        
        hoja["B19"] = capitalizar(informacion_adicional.jornada)
        hoja["H17"] = capitalizar(informacion_adicional.modalidad_formacion)

        datos_fechas = {
            "C19": fecha_inicio,
            "D19": fecha_fin,
            "H19": fecha_fin,
            "B12": fecha_actual,
            "G19": fecha_productiva  
        }
        
        # Aplicar todo de una vez
        font_style = Font(size=12, bold=True)
        for celda, valor in datos_fechas.items():
            if valor:  # Si hay fecha válida
                hoja[celda].value = valor
                hoja[celda].number_format = "DD-MM-YYYY"  # <-- Forzar formato de fecha
            else:
                hoja[celda].value = "N/A"
            hoja[celda].font = font_style

        hoja["D22"] = "X"
        hoja["F30"] = f"{ap.nombre} {ap.apellido}"

        # Firma (imagen)
        if imagen_path:
            try:
                img = OpenpyxlImage(imagen_path)
                img.width = 120
                img.height = 50
                hoja.add_image(img, "G30")  # Aquí va como string, no lista
            except Exception as e:
                print(f"Error insertando imagen en hoja individual: {e}")

    def generar_f165_grupal(self, ficha, aprendices, imagenes_procesadas, request, usuario_gene, informacion_adicional):
        """
        Función pública que prepara y genera el formato F165 grupal.
        """

        wb_copia = load_workbook(self.plantilla_grupal_wb)


        # 2. Llama a tu función de llenado, pero pasándole la COPIA
        self._llenar_F165_grupal(
            wb_copia,
            ficha,
            aprendices,
            imagenes_procesadas,
            request,
            usuario_gene,
            informacion_adicional
        )
        return wb_copia

    def generar_f165_individual(self, ficha, aprendices, imagenes_procesadas, request, usuario_gene, informacion_adicional):
        """
        Función pública que prepara y genera el formato F165 individual.
        """
        # 1. Crea una copia profunda para no modificar la plantilla original en memoria
        wb_copia = load_workbook(self.plantilla_individual_wb)
        print("Hojas en plantilla individual:", wb_copia.sheetnames)

        # 2. Llama a tu función de llenado, pero pasándole la COPIA
        self._llenar_F165_individual(
            wb_copia,
            ficha,
            aprendices,
            imagenes_procesadas,
            request,
            usuario_gene,
            informacion_adicional
        )
        return wb_copia
    
    def crear_y_guardar_formato_f165(self, db:Session, request, modalidad:str, aprendices:list,usuario_gene,informacion_adicional,imagenes_procesadas):
        ficha = self._validar_y_obtener_ficha(request.ficha,db)

        wb = None
        nombre_original = None
        aprendiz_documento = None 
        ap = aprendices[0]
        
        if modalidad == "grupal":
            wb = self.generar_f165_grupal(ficha, aprendices, imagenes_procesadas, request, usuario_gene, informacion_adicional)
            nombre_original = f"F165_{request.ficha}_{request.modalidad}"
            print("este es el nombre del archivo", nombre_original)
            
        elif modalidad == "individual":
            wb = self.generar_f165_individual(ficha, aprendices, imagenes_procesadas, request, usuario_gene, informacion_adicional)

            if aprendices and len(aprendices) > 0:
                aprendiz_documento = ap.documento  # 👈 sacar el documento
                print("documento aprendiz", ap.documento)
                nombre_original = f"F165_{request.ficha}_{modalidad}_{aprendiz_documento}"
                print("este es el nombre del archivo", nombre_original)

        else:
            raise Exception("Modalidad no válida")
        
        stream = BytesIO()
        
        # Verificar que no hay celdas problemáticas
        print(f"Worksheets en wb: {[ws.title for ws in wb.worksheets]}")
        print(f"Hoja activa: {wb.active.title}")
        
        print("Hojas disponibles:", wb.sheetnames)

        if modalidad == "grupal":
            hoja = wb["Selección formato - Grupal"]
        elif modalidad == "individual":
            hoja = wb["Selección Modificación Indiv"]
        else:
            raise Exception("Modalidad no válida")

        try:
            wb.save(stream)  
        except Exception as e:
            print(f"Error al guardar workbook: {e}")
            print(f"Tipo de error: {type(e)}")
            import traceback
            traceback.print_exc()  # Esto te dará más detalles del error
            raise

        stream.seek(0)
        contenido_bytes = stream.read()
        print(f"Contenido leído: {len(contenido_bytes)} bytes")


        print("Llamando a guardar_archivo_seguro...")
        archivo_db = self.guardar_archivo_seguro(
            contenido=contenido_bytes,
            nombre_original=nombre_original,
            ficha=request.ficha,
            modalidad=modalidad,
            cantidad_aprendices=len(aprendices),
            usuario_id=usuario_gene.id,
            aprendiz_documento=aprendiz_documento
        )

        print("Archivo guardado en memoria correctamente")

        print("Guardando en base de datos...")
        try:
            db.add(archivo_db)
            db.commit()
            db.refresh(archivo_db)
            print("Guardado en BD correctamente")
        except Exception as e:
            db.rollback()
            raise Exception(f"Error al guardar en la base de datos: {str(e)}") from e

        ruta_completa = self.base_path / archivo_db.ruta_archivo
        print(f"Ruta completa: {ruta_completa}")
        return archivo_db, ruta_completa
